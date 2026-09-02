#matcher.py
"""
matcher.py
Mesin pencocokan uraian surat -> Kode Klasifikasi PERGUB, memakai kombinasi
kamus istilah (dari kamus_istilah_pencocokan.xlsx) + TF-IDF, sepenuhnya offline & gratis.
"""
import json
import re
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity


class KlasifikasiMatcher:
    def __init__(self, master_json_jra, master_json_skkd, kamus_xlsx_path):
        with open(master_json_jra, encoding="utf-8") as f:
            jra = json.load(f)
        with open(master_json_skkd, encoding="utf-8") as f:
            skkd = json.load(f)
        self.jra_map = {r["kode_klasifikasi"]: r for r in jra}
        self.skkd_map = {r["kode_klasifikasi"]: r for r in skkd}

        self.sub_kegiatan_map = {}   # label dropdown -> kode klasifikasi
        self.alias = {}              # kode -> string kata kunci tambahan
        self._load_kamus(kamus_xlsx_path)

        self._build_corpus()
        self._build_sub_kegiatan_terms()

    def _build_sub_kegiatan_terms(self):
        """Gabungkan label sub kegiatan + kata kunci dari kode terkait -> daftar term per label."""
        self.sub_kegiatan_terms = {}
        for label, kode in self.sub_kegiatan_map.items():
            terms = set()
            for w in re.split(r"\s+", label.lower()):
                w = w.strip(",.")
                if len(w) > 3:
                    terms.add(w)
            alias_text = self.alias.get(kode, "")
            for term in alias_text.split(","):
                term = term.strip().lower()
                if term:
                    terms.add(term)
            self.sub_kegiatan_terms[label] = terms

    @staticmethod
    def _retensi_tahun(value):
        match = re.search(r"(\d+)\s*tahun", str(value or ""), re.IGNORECASE)
        return f"{match.group(1)} TAHUN" if match else ""

    @staticmethod
    def _normalize_text_upper(value):
        return str(value or "").strip().upper()

    @staticmethod
    def _normalize_skkd_keamanan(value):
        v = str(value or "").strip().upper()
        mapping = {
            "BIAASA": "TERBUKA",
            "BIASA": "TERBUKA",
            "TERBUKA": "TERBUKA",
            "TERBATAS": "TERTUTUP",
            "TERTUTUP": "TERTUTUP",
            "TERBATASI": "TERTUTUP",
        }
        return mapping.get(v, v if v else "")

    def classify_sub_kegiatan(self, text, top_n=3):
        """
        Tebak Sub Kegiatan (dari 10 pilihan tetap) berdasarkan kecocokan kata kunci.
        Kembalikan list (label, skor) terurut, skor = total kata dari term yang cocok
        (term multi-kata dihitung lebih berat karena lebih spesifik).
        """
        text_lower = text.lower()
        skor_label = []
        for label, terms in self.sub_kegiatan_terms.items():
            skor = 0
            for term in terms:
                if term in text_lower:
                    skor += len(term.split())
            skor_label.append((label, skor))
        skor_label.sort(key=lambda x: -x[1])
        return skor_label[:top_n]

    def _load_kamus(self, path):
        wb = load_workbook(path, data_only=True)
        ws1 = wb["Peta 10 Sub Kegiatan"]
        for row in ws1.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            label, kode = row[0], row[1]
            if kode and kode != "BELUM KETEMU":
                self.sub_kegiatan_map[label.strip()] = kode.strip()

        ws2 = wb["Kamus Kata Kunci (Desa)"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            kode, _, kata_kunci = row[0], row[1], row[2]
            self.alias[kode.strip()] = kata_kunci or ""

    def _build_corpus(self):
        all_kode = sorted(set(self.jra_map) | set(self.skkd_map))
        self.kodes, docs = [], []
        for k in all_kode:
            j, s = self.jra_map.get(k), self.skkd_map.get(k)
            desc = (j or s)["jenis_series_arsip"]
            enriched = desc + " " + self.alias.get(k, "")
            self.kodes.append(k)
            docs.append(enriched)
        self.docs_original = {k: (self.jra_map.get(k) or self.skkd_map.get(k))["jenis_series_arsip"] for k in all_kode}
        self.vectorizer = TfidfVectorizer(lowercase=True, ngram_range=(1, 2), min_df=1)
        self.tfidf_matrix = self.vectorizer.fit_transform(docs)

    def get_retensi_skkd(self, kode):
        j, s = self.jra_map.get(kode), self.skkd_map.get(kode)
        return {
            "kode_klasifikasi": kode,
            "jenis_series_arsip": self.docs_original.get(kode, ""),
            "retensi_aktif": self._retensi_tahun(j["retensi_aktif"] if j else ""),
            "retensi_inaktif": self._retensi_tahun(j["retensi_inaktif"] if j else ""),
            "keterangan": self._normalize_text_upper(j["keterangan"] if j else ""),
            "klasifikasi_keamanan": self._normalize_skkd_keamanan(s["klasifikasi_keamanan"] if s else ""),
            "hak_akses": s["hak_akses"] if s else "",
            "unit_pengolah": s["unit_pengolah"] if s else "",
        }

    def cari_kandidat(self, uraian_surat, sub_kegiatan=None, top_n=3):
        """
        Kembalikan top_n kandidat kode klasifikasi.
        Kalau sub_kegiatan dikenal & sudah ada pemetaan pasti -> kode itu selalu
        jadi kandidat #1 (skor 1.0, sumber='pemetaan sub kegiatan'), sisanya
        pelengkap dari TF-IDF supaya pegawai tetap punya alternatif.
        """
        hasil = []
        kode_pasti = self.sub_kegiatan_map.get((sub_kegiatan or "").strip())
        if kode_pasti:
            hasil.append({**self.get_retensi_skkd(kode_pasti), "skor": 1.0, "sumber": "pemetaan sub kegiatan"})

        qv = self.vectorizer.transform([uraian_surat])
        sims = cosine_similarity(qv, self.tfidf_matrix).flatten()
        idx_sorted = sims.argsort()[::-1]
        for i in idx_sorted:
            kode = self.kodes[i]
            if kode_pasti and kode == kode_pasti:
                continue
            if sims[i] <= 0:
                break
            hasil.append({**self.get_retensi_skkd(kode), "skor": round(float(sims[i]), 3), "sumber": "kemiripan teks"})
            if len(hasil) >= top_n:
                break
        return hasil[:top_n]


if __name__ == "__main__":
    m = KlasifikasiMatcher("lampiran_ii_jra.json", "lampiran_iii_skkd.json",
                            "/mnt/user-data/outputs/kamus_istilah_pencocokan.xlsx")

    tests = [
        ("Permohonan pembinaan keluarga BPD", None),
        ("Dukungan pelaksaan pelantikan pengurus PPDI Kaltim", None),
        ("Undangan sosialisasi tahapan pemilihan kepala desa serentak", "Fasilitasi PILKADES"),
        ("Permohonan Narasumber untuk kegiatan pembinaan aparatur desa", "Pembinaan Aparatur Pemerintahan Desa"),
    ]
    for uraian, sub in tests:
        print(f"URAIAN: {uraian}")
        print(f"SUB KEGIATAN (dropdown): {sub}")
        for kand in m.cari_kandidat(uraian, sub_kegiatan=sub, top_n=3):
            print(f"   [{kand['skor']}] ({kand['sumber']}) {kand['kode_klasifikasi']} - {kand['jenis_series_arsip'][:60]}")
            print(f"        Retensi: aktif={kand['retensi_aktif']}, inaktif={kand['retensi_inaktif']}, ket={kand['keterangan']}")
            print(f"        SKKD: {kand['klasifikasi_keamanan']}")
        print()
